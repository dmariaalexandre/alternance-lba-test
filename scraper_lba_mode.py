"""
Scraper LBA (La Bonne Alternance) â Mode / Ãle-de-France â TEST
================================================================
Sources  : peJobs (France Travail) + lbaJobs (offres directes LBA)
Exclut   : lbaCompanies (entreprises suggÃ©rÃ©es) + matchas (formations)
Zone     : Paris + 30 km  (couvre toute la petite couronne IDF)
Filtre   : offres des 14 derniers jours
Output   : lba_mode_YYYY-MM-DD.xlsx  +  lba_mode_YYYY-MM-DD.csv

Secrets requis (GitHub Secrets) :
  - LBA_API_TOKEN  â token api.apprentissage.beta.gouv.fr

Mode debug : LBA_DEBUG=true  â affiche le JSON brut de la premiÃ¨re page
"""

import csv
import hashlib
import json
import os
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# CONFIG
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

LBA_API_TOKEN = os.getenv("LBA_API_TOKEN", "")
LBA_JOBS_URL  = "https://labonnealternance.apprentissage.beta.gouv.fr/api/v1/jobs"

# Codes ROME mode (12 codes, B1806 exclu â tapisserie â  mode)
ROMES_MODE = [
    "B1801",  # Chapellerie / Modiste
    "B1803",  # VÃªtements sur mesure / petite sÃ©rie
    "B1805",  # Stylisme
    "B1808",  # Confection, production en sÃ©rie
    "B1809",  # Couture flou
    "B1813",  # Maroquinerie et gainerie
    "H1205",  # Ãtudes-modÃ¨les matÃ©riaux souples (modÃ©liste industriel)
    "H2401",  # Assemblage-montage cuirs/ peaux
    "H2402",  # Assemblage-montage vÃªtements / textiles
    "H2411",  # Montage prototype cuir / matÃ©riaux souples
    "H2412",  # Patronnage-gradation
    "D1214",  # Vente en habillement et accessoires de la personne
]

# Paris centre â 30 km couvre 92/93/94 + majoritÃ© de la grande couronne
PARIS_LAT = 48.8534
PARIS_LON = 2.3488
RADIUS_KM = 30

DAYS_BACK  = 14     # filtre : offres des N derniers jours
PAGE_LIMIT = 50     # items par page (max LBA)

OUTPUT_DIR = Path(__file__).parent
TODAY      = date.today().isoformat()
CUTOFF     = datetime.now(tz=timezone.utc) - timedelta(days=DAYS_BACK)

# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# FORMAT EXCEL (compatible avec la production)
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

# (label affichÃ©, largeur colonne, clÃ© interne dans l'offre)
COLUMNS = [
    ("Statut",           14, "statut"),
    ("Titre du poste",   38, "titre"),
    ("Entreprise",       22, "entreprise"),
    ("Ville",            20, "ville"),
    ("RÃ©gion",           18, "region"),
    ("DÃ©partement",      18, "departement"),
    ("CatÃ©gorie",        16, "categorie"),
    ("Contrat",          14, "contrat"),
    ("Type emploi",      14, "type_emploi"),
    ("ExpÃ©rience",       16, "experience"),
    ("Date dÃ©but",       14, "date_debut"),
    ("Date publication", 16, "date_publication"),
    ("Description",      70, "description"),
    ("Profil recherchÃ©", 60, "profil_recherche"),
    ("Contact nom",      22, "contact_nom"),
    ("Contact tÃ©l",      16, "contact_tel"),
    ("Contact email",    30, "contact_email"),   # enrichissement LBA
    ("Site internet",    35, "site_internet"),
    ("Lien candidature", 45, "lien_candidature"),
    ("Source",           20, "source"),
    ("Lien offre",       50, "lien"),
    ("Notes",            30, "notes"),
    ("ID",               18, "id"),
    ("Date scraping",    14, "date_scraping"),
    ("Code ROME",        12, "rome_debug"),       # debug LBA
]

HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

SOURCE_COLORS = {
    "LBA - France Travail": "E3F2FD",   # bleu clair (peJobs)
    "LBA - Direct":         "E8F5E9",   # vert clair (lbaJobs)
}

WRAP_COLS = {13, 14}  # indices (1-based) des colonnes Description + Profil
LINK_COLS = {         # col_idx â clÃ© interne pour hyperlien
    17: "site_internet",
    19: "lien_candidature",
    21: "lien",
}


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# API LBA
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def _headers() -> dict:
    h = {"Accept": "application/json"}
    if LBA_API_TOKEN:
        h["Authorization"] = f"Bearer {LBA_API_TOKEN}"
    return h


def fetch_page(romes_str: str, page: int) -> dict:
    """Appelle l'endpoint /api/v1/jobs et retourne le JSON brut."""
    params = {
        "romes":     romes_str,
        "longitude": PARIS_LON,
        "latitude":  PARIS_LAT,
        "radius":    RADIUS_KM,
        "caller":    "lba-mode-test-scraper",
        "limit":     PAGE_LIMIT,
        "page":      page,
    }
    try:
        r = requests.get(LBA_JOBS_URL, headers=_headers(), params=params, timeout=30)
        r.raise_for_status()
        return r.json()
    except requests.HTTPError as e:
        print(f"  â ï¸  HTTP {e.response.status_code} page {page}")
        return {}
    except Exception as e:
        print(f"  â ï¸  Erreur rÃ©seau page {page} : {e}")
        return {}


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# PARSING
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def _parse_date(raw: str | None) -> datetime | None:
    if not raw:
        return None
    for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ",
                "%Y-%m-%dT%H:%M:%S%z", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(raw[:26], fmt)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt
        except ValueError:
            continue
    return None


def is_recent(date_raw: str | None) -> bool:
    """True si la date est dans les DAYS_BACK derniers jours (ou absente â inclus)."""
    if not date_raw:
        return True
    dt = _parse_date(date_raw)
    return dt >= CUTOFF if dt else True


def _get(obj: dict, *keys, default="") -> str:
    """Cherche les clÃ©s alternatives dans un dict, retourne le premier non vide."""
    for k in keys:
        v = obj.get(k)
        if v and str(v).strip():
            return str(v).strip()
    return default


def parse_offer(raw: dict, source_label: str) -> dict | None:
    """
    Normalise une offre LBA vers le schÃ©ma de production.
    Retourne None si l'offre est trop ancienne ou sans titre.

    Structure attendue (LBA v1 â peut varier selon la version) :
      raw.job         â mÃ©tadonnÃ©es du poste
      raw.company      â infos entreprise
      raw.place         â localisation
      raw.apply        â comment postuler (url, phone, email)
      raw.contact       â contact RH (lbaJobs seulement)
    """
    job     = raw.get("job",     {}) or {}
    company = raw.get("company", {}) or {}
    place   = raw.get("place",   {}) or {}
    apply   = raw.get("apply",   {}) or {}
    contact = raw.get("contact", {}) or {}

    # ââ Date de publication ââââââââââââââââââââââââââââââââââââââââââââââââââââ
    date_raw = _get(job,  "dateCreation", "date_creation", "createdAt") \
            or _get(raw,  "createdAt", "dateCreation")
    if not is_recent(date_raw):
        return None
    date_pub = date_raw[:10] if date_raw else ""

    # ââ Titre ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    titre = _get(job, "title", "intitule") or _get(raw, "title", "intitule")
    if not titre:
        return None

    # ââ Entreprise âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    entreprise = _get(company, "name", "enseigne", "raison_sociale") \
              or _get(raw.get("company", {}), "name")

    # ââ Localisation ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    ville       = _get(place, "city", "ville") or _get(job, "locationDisplay", "location")
    region      = _get(place, "region") or "Ãle-de-France"
    departement = _get(place, "departement", "department")

    # ââ Contrat ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    contrat = _get(job, "contractType", "contract_type", "typeContrat") or "Alternance"

    # ââ Description ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    description = (_get(job, "description") or _get(raw, "description"))[:3000]

    # ââ Contact ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    phone = (_get(apply, "phone", "phoneNumber")
             or _get(contact, "phone", "telephone"))
    email = (_get(apply, "email")
             or _get(contact, "email"))

    first  = _get(contact, "firstName", "prenom")
    last   = _get(contact, "lastName",  "nom")
    cname  = _get(apply, "name") or _get(contact, "name") or \
             (f"{first} {last}".strip() if first or last else "")

    # ââ Liens ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    apply_url = _get(apply, "url", "urlPostuler")
    site      = _get(company, "url", "website", "site")

    # ââ Code ROME (debug) âââââââââââââââââââââââââââââââââââââââââââââââââââââ
    rome = _get(job, "rome", "romeCode") or _get(raw, "romeLabel", "rome")

    # ââ ID unique ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
    uid_src = apply_url or _get(job, "id") or _get(raw, "id") or f"{titre}|{entreprise}"
    offer_id = hashlib.sha256(uid_src.encode()).hexdigest()[:16]

    return {
        "statut":           "Nouveau",
        "titre":            titre,
        "entreprise":       entreprise,
        "ville":            ville,
        "region":           region,
        "departement":      departement,
        "categorie":        "",
        "contrat":          contrat,
        "type_emploi":      "",
        "experience":       "",
        "date_debut":       "",
        "date_publication": date_pub,
        "description":      description,
        "profil_recherche": "",
        "contact_nom":      cname,
        "contact_tel":      phone,
        "contact_email":    email,
        "site_internet":    site,
        "lien_candidature": apply_url,
        "source":           source_label,
        "lien":             apply_url,
        "notes":            "",
        "id":               offer_id,
        "date_scraping":    TODAY,
        "rome_debug":       rome,
    }


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# SCRAPE
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def scrape_lba(debug: bool = False) -> list[dict]:
    """
    Pagine l'API LBA jusqu'Ã  Ã©puisement ou jusqu'Ã  ce que toutes les offres
    soient plus vieilles que DAYS_BACK.
    """
    romes_str = ",".join(ROMES_MODE)
    all_offers: list[dict] = []
    total_raw  = 0

    print(f"\nð LBA â {len(ROMES_MODE)} codes ROME mode Â· Paris +{RADIUS_KM} km Â· {DAYS_BACK} derniers jours")
    print(f"   Codes : {romes_str}")
    print(f"   Token : {'â prÃ©sent' if LBA_API_TOKEN else 'â ï¸  absent (appel sans auth)'}")

    page = 0
    while True:
        print(f"  â Page {page}...", end=" ", flush=True)
        data = fetch_page(romes_str, page)

        if not data:
            print("erreur ou rÃ©ponse vide.")
            break

        # ââ Debug : structure JSON brute sur la page 0 âââââââââââââââââââââ
        if debug and page == 0:
            print("\n\n  ð DEBUG â structure rÃ©ponse page 0 :")
            summary = {
                k: (f"{len(v.get('results', []))} items" if isinstance(v, dict) else type(v).__name__)
                for k, v in data.items()
            }
            print(json.dumps(summary, indent=2))

            for key in ("peJobs", "lbaJobs"):
                results = data.get(key, {}).get("results", [])
                if results:
                    print(f"\n  ð Premier item de '{key}' :")
                    print(json.dumps(results[0], indent=2, ensure_ascii=False)[:2000])
            print()

        pe_results  = (data.get("peJobs",  None) or {}).get("results",  []) or []
        lba_results = (data.get("lbaJobs", None) or {}).get("results",  []) or []
        # lbaCompanies et matchas â ignorÃ©s

        page_raw = len(pe_results) + len(lba_results)
        if page_raw == 0:
            print("aucun rÃ©sultat, arrÃªt.")
            break

        total_raw += page_raw
        kept = too_old = skipped = 0

        for raw in pe_results:
            offer = parse_offer(raw, "LBA - France Travail")
            if offer:
                all_offers.append(offer)
                kept += 1
            else:
                too_old += 1

        for raw in lba_results:
            offer = parse_offer(raw, "LBA - Direct")
            if offer:
                all_offers.append(offer)
                kept += 1
            else:
                too_old += 1

        print(f"{kept} rÃ©centes / {page_raw} ({too_old} trop anciennes, {skipped} sans titre)")

        # Toutes les offres de la page sont trop anciennes â inutile de paginer
        if too_old == page_raw and page > 0:
            print("  â¹  Toutes > 14 jours â arrÃªt pagination.")
            break

        # DerniÃ¨re page si on reÃ§oit moins que PAGE_LIMIT
        if page_raw < PAGE_LIMIT:
            break

        page += 1

    print(f"\nâ RÃ©cupÃ©ration terminÃ©e : {len(all_offers)} offres rÃ©centes / {total_raw} analysÃ©es")
    return all_offers


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# DÃDUPLICATION
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def deduplicate(offers: list[dict]) -> list[dict]:
    seen: set[tuple] = set()
    unique: list[dict] = []
    for o in offers:
        key = (o["titre"].lower().strip(), o["entreprise"].lower().strip())
        if key not in seen:
            seen.add(key)
            unique.append(o)
    removed = len(offers) - len(unique)
    if removed:
        print(f"ð DÃ©duplication : {removed} doublons supprimÃ©s â {len(unique)} offres uniques")
    return unique


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# EXPORT EXCEL
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def export_excel(offers: list[dict], filepath: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "LBA Mode IDF"

    # En-tÃªtes
    for col_idx, (label, width, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    # DonnÃ©es
    for row_idx, offer in enumerate(offers, start=2):
        src_color = SOURCE_COLORS.get(offer.get("source", ""), "FFFFFF")
        row_fill  = PatternFill("solid", fgColor=src_color)

        for col_idx, (_, _, key) in enumerate(COLUMNS, start=1):
            val  = offer.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(
                wrap_text=(col_idx in WRAP_COLS),
                vertical="top",
            )

        # Liens cliquables
        for col_idx, key in LINK_COLS.items():
            lnk = offer.get(key, "")
            if lnk and lnk.startswith("http"):
                lc = ws.cell(row=row_idx, column=col_idx)
                lc.hyperlink = lnk
                lc.font = Font(color="0563C1", underline="single")

    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"ð¾ Excel : {filepath.name}  ({filepath.stat().st_size // 1024} ko)")


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# EXPORT CSV
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def export_csv(offers: list[dict], filepath: Path) -> None:
    if not offers:
        return
    fieldnames = [col[0] for col in COLUMNS]
    key_map    = {col[0]: col[2] for col in COLUMNS}

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for o in offers:
            writer.writerow({col: o.get(key_map[col], "") for col in fieldnames})

    print(f"ð¾ CSV   : {filepath.name}  ({filepath.stat().st_size // 1024} ko)")


# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
# MAIN
# ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ

def main() -> None:
    debug = os.getenv("LBA_DEBUG", "").lower() in ("1", "true", "yes")

    print(f"ð LBA Mode Scraper â {TODAY}")
    print("=" * 60)

    offers = scrape_lba(debug=debug)
    offers = deduplicate(offers)

    if not offers:
        print("\nâ ï¸  Aucune offre rÃ©cupÃ©rÃ©e.")
        print("   â VÃ©rifier : LBA_API_TOKEN valide ? Connexion rÃ©seau ?")
        print("   â Relancer avec LBA_DEBUG=true pour voir la structure API.")
        return

    # Statistiques rapides
    with_phone = sum(1 for o in offers if o.get("contact_tel"))
    with_email = sum(1 for o in offers if o.get("contact_email"))
    ft_count   = sum(1 for o in offers if "France Travail" in o.get("source", ""))
    lba_count  = sum(1 for o in offers if "Direct" in o.get("source", ""))

    print(f"\nð RÃ©sultats â {TODAY}")
    print(f"   Total offres    : {len(offers)}")
    print(f"   France Travail  : {ft_count}  (peJobs)")
    print(f"   LBA Direct      : {lba_count}  (lbaJobs)")
    print(f"   Avec tÃ©lÃ©phone  : {with_phone}")
    print(f"   Avec email      : {with_email}")
    print()

    xlsx_path = OUTPUT_DIR / f"lba_mode_{TODAY}.xlsx"
    csv_path  = OUTPUT_DIR / f"lba_mode_{TODAY}.csv"

    export_excel(offers, xlsx_path)
    export_csv(offers,   csv_path)

    print(f"\nâ Done â {len(offers)} offres Â· {DAYS_BACK}j Â· Paris +{RADIUS_KM}km Â· {len(ROMES_MODE)} codes ROME")


if __name__ == "__main__":
    main()
